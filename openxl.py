#!/usr/bin/python
# coding:utf-8
import sys
reload(sys)
sys.setdefaultencoding('utf8')

from openpyxl import load_workbook
from collections import OrderedDict
import json
import copy


def write_Conf(path,data):
    with open(path, 'w') as f:
        f.write(data + "\n")

res_json = {
    'd':'刘焕',
    'b':'ddd换aa'
}

# write_Conf("json1.json", json.dumps(res_json,ensure_ascii=False))


wb = load_workbook('item.xlsx',data_only=True)


print wb.get_sheet_names()

sheet = wb.get_sheet_by_name('item')

print 'max row',sheet.max_row
print 'max column',sheet.max_column

res_json = OrderedDict()

row_list = list(sheet.rows)
# for i in row_list[0]:
#     print i.value


# for r in xrange(sheet.max_row -2):
keyDict = OrderedDict()
for cell in row_list[1]:
    keyDict.setdefault(str(cell.value))

# for i in row_list[2:]:
#     print i
key_id = 'img'
for row in row_list[2:]:

    idx = 0
    key = None
    keyDictCopy = copy.deepcopy(keyDict)
    for k,v in keyDict.iteritems():
        if k == key_id:
            key = row[idx].value
        value = row[idx].value

        # 不填值 将不会输出
        if value is None:
            del keyDictCopy[k]
            idx += 1
            continue

        # 字符串
        if row[idx].data_type == 's':
            # 处理 {} []
            # value = value.encode('utf8')

            if value.startswith('{') or value.startswith('['):
                value = json.loads(value)
            else:
                value = value.encode('utf8')

        keyDictCopy[k] = value
        idx += 1

    res_json.update({key: keyDictCopy})


write_Conf("json.json", json.dumps(res_json,ensure_ascii=False))


print sheet.columns
print sheet.rows


print sheet['B2']




